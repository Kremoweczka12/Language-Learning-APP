
from dataclasses import make_dataclass

from openpyxl import load_workbook

from parsers.BaseParser import BaseParser
from utils.global_access_classes import Const, Config


class ExcelGrandParser(BaseParser):

    def _get_numbers_for_headers(self, worksheet):
        if hasattr(self.config, "interesting_columns") and self.config.interesting_columns is not None \
                and self.config.interesting_columns != ['']:
            for i, cell in enumerate(worksheet['1']):
                value = str(cell.value)
                if value in self.config.interesting_columns:
                    self.headers_to_numbers[self.remove_wrong_letters(value)] = i
        else:
            self.headers_to_numbers = {self.remove_wrong_letters(str(cell.value)): i for i, cell in
                                       enumerate(worksheet['1'])}
        self.headers = [(self.remove_wrong_letters(str(header)), str) for header in self.headers_to_numbers.keys()]

    def __init__(self, config: Config):
        super().__init__()
        self.headers = []
        self.headers_to_numbers = {}
        self.config = config
        self.dataset_name = config.absolute_path_to_file

        workbook = load_workbook(config.absolute_path_to_file)

        worksheet = workbook.worksheets[0]
        self._get_numbers_for_headers(worksheet)
        self.RecordClass = make_dataclass("Record", self.headers)  # , headers
        self.all_records = self._add_records(worksheet)

    def _add_records(self, worksheet):
        records = []
        for row in worksheet.iter_rows(2, ):
            tab = (cell.value for i, cell in enumerate(row) if i in self.headers_to_numbers.values())
            record = self.RecordClass(*tab)
            record.ID = next(Const.GRAND_ITER)
            records.append(record)
        return records
